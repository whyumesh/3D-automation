import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy as copy_style
import warnings

# Suppress FutureWarning for groupby operations
warnings.filterwarnings('ignore', category=FutureWarning, module='pandas')

def create_zbm_hierarchical_reports():
    """
    Create separate ZBM reports showing ABM hierarchy with perfect tallies
    Each ZBM gets a report showing all ABMs under them
    """
    
    print("🔄 Starting ZBM Hierarchical Reports Creation...")
    
    # Read master tracker data from Excel file
    print("📖 Reading Sample Master Tracker.xlsx...")
    try:
        df = pd.read_excel('Sample Master Tracker.xlsx')
        print(f"✅ Successfully loaded {len(df)} records from Sample Master Tracker.xlsx")
    except Exception as e:
        print(f"❌ Error reading Sample Master Tracker.xlsx: {e}")
        return
    
    # Using raw Request Status field from Sample Master Tracker (no Final Status computation needed)
    print("📊 Using raw Request Status field from Sample Master Tracker for accurate counts...")
    
    # Clean and prepare data
    print("🧹 Cleaning and preparing data...")
    
    # Ensure required columns exist
    required_columns = ['ZBM Terr Code', 'ZBM Name', 'ZBM EMAIL_ID',
                        'ABM Terr Code', 'ABM Name', 'ABM EMAIL_ID',
                        'TBM HQ', 'TBM EMAIL_ID',
                        'Doctor: Customer Code', 'Assigned Request Ids', 'Request Status', 'Rto Reason']
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        print(f"❌ Missing required columns in Sample Master Tracker.xlsx: {missing}")
        return

    # Remove rows where key fields are null or empty
    df = df.dropna(subset=['ZBM Terr Code', 'ZBM Name', 'ABM Terr Code', 'ABM Name', 'TBM HQ'])
    df = df[df['ZBM Terr Code'].astype(str).str.strip() != '']
    df = df[df['ABM Terr Code'].astype(str).str.strip() != '']
    df = df[df['TBM HQ'].astype(str).str.strip() != '']

    # Filter for ZBM codes that start with "ZN" (only restriction needed)
    df = df[df['ZBM Terr Code'].astype(str).str.startswith('ZN')]
    print(f"📊 After cleaning and ZBM filtering: {len(df)} records remaining")
    print(f"📊 Processing all ZBM codes starting with 'ZN' - no geographic restrictions")

    # Compute Final Answer per unique request id using rules from logic.xlsx
    print("🧠 Computing final status per unique Request Id using rules...")
    try:
        xls_rules = pd.ExcelFile('logic.xlsx')
        sheet2 = pd.read_excel(xls_rules, 'Sheet2')

        def normalize(text):
            return str(text).strip().casefold()

        rules = {}
        for _, row in sheet2.iterrows():
            statuses = [normalize(s) for s in row.drop('Final Answer').dropna().tolist()]
            statuses = tuple(sorted(set(statuses)))
            rules[statuses] = row['Final Answer']

        # Group statuses by request id from master data
        grouped = df.groupby('Assigned Request Ids')['Request Status'].apply(list).reset_index()

        def get_final_answer(status_list):
            key = tuple(sorted(set(normalize(s) for s in status_list)))
            return rules.get(key, '❌ No matching rule')

        grouped['Request Status'] = grouped['Request Status'].apply(lambda lst: sorted(set(lst), key=str))
        grouped['Final Answer'] = grouped['Request Status'].apply(get_final_answer)

        def has_action_pending(status_list):
            target = 'action pending / in process'
            return any(normalize(s) == target for s in status_list)
        grouped['Has D Pending'] = grouped['Request Status'].apply(has_action_pending)

        # Merge Final Answer back to main dataframe
        df = df.merge(grouped[['Assigned Request Ids', 'Final Answer', 'Has D Pending']], on='Assigned Request Ids', how='left')
    except Exception as e:
        print(f"❌ Error computing final status from logic.xlsx: {e}")
        return
    
    # Get unique ZBMs
    zbms = df[['ZBM Terr Code', 'ZBM Name', 'ZBM EMAIL_ID']].drop_duplicates().sort_values('ZBM Terr Code')
    print(f"📋 Found {len(zbms)} unique ZBMs")
    
    # Debug: Show all ZBMs and their ABMs
    print("\n🔍 ZBM-ABM Mapping:")
    for _, zbm_row in zbms.iterrows():
        zbm_code = zbm_row['ZBM Terr Code']
        zbm_name = zbm_row['ZBM Name']
        zbm_data_temp = df[df['ZBM Terr Code'] == zbm_code]
        abms_temp = zbm_data_temp[['ABM Terr Code', 'ABM Name']].drop_duplicates()
        print(f"   {zbm_code} ({zbm_name}): {len(abms_temp)} ABMs")
        for _, abm_row in abms_temp.iterrows():
            print(f"      - {abm_row['ABM Terr Code']}: {abm_row['ABM Name']}")
    
    # Create output directory
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = f"ZBM_Reports_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Created output directory: {output_dir}")
    
    # Define status categories for calculations
    status_categories = {
        'out_of_stock_on_hold': ['Out of stock', 'On hold', 'Not permitted'],
        'request_raised': ['Request Raised'],
        'delivered_return_action_pending': ['Delivered', 'Return', 'Action pending / In Process', 'Dispatched & In Transit', 'Dispatch Pending'],
        'action_pending': ['Action pending / In Process'],
        'dispatch_pending': ['Dispatch Pending'],
        'delivered': ['Delivered'],
        'dispatched_in_transit': ['Dispatched & In Transit'],
        'rto': ['RTO']
    }
    
    # Process each ZBM
    for _, zbm_row in zbms.iterrows():
        zbm_code = zbm_row['ZBM Terr Code']
        zbm_name = zbm_row['ZBM Name']
        zbm_email = zbm_row['ZBM EMAIL_ID']
        
        print(f"\n🔄 Processing ZBM: {zbm_code} - {zbm_name}")
        
        # Filter data for this ZBM
        zbm_data = df[df['ZBM Terr Code'] == zbm_code].copy()
        
        if len(zbm_data) == 0:
            print(f"⚠️ No data found for ZBM: {zbm_code}")
            continue
        
        # Get unique ABMs under this ZBM (properly filtered and aggregated)
        # Group by ABM Terr Code and ABM Name to avoid duplicates
        abms = zbm_data.groupby(['ABM Terr Code', 'ABM Name']).agg({
            'ABM EMAIL_ID': 'first',
            'TBM HQ': 'first',
            'ABM HQ': 'first' if 'ABM HQ' in zbm_data.columns else lambda x: None
        }).reset_index()
        
        # Sort by ABM Terr Code
        abms = abms.sort_values('ABM Terr Code')
        print(f"   📊 Found {len(abms)} ABMs under this ZBM")
        
        # Debug: Show ABM names to verify filtering
        abm_names = abms['ABM Name'].tolist()
        print(f"   📋 ABMs under {zbm_code}: {abm_names}")
        
        # Validate that Rashmi Sharma appears only in her ZBM
        if 'Rashmi Sharma' in abm_names:
            print(f"   ⚠️ Rashmi Sharma found in {zbm_code} - {zbm_name}")
        
        # Create summary data for this ZBM
        summary_data = []
        
        for _, abm_row in abms.iterrows():
            abm_code = abm_row['ABM Terr Code']
            abm_name = abm_row['ABM Name']
            abm_email = abm_row['ABM EMAIL_ID']
            tbm_hq = abm_row['TBM HQ']
            
            # Filter data for this specific ABM (by both code and name to ensure uniqueness)
            abm_data = zbm_data[(zbm_data['ABM Terr Code'] == abm_code) & (zbm_data['ABM Name'] == abm_name)]
            
            print(f"      Processing {abm_name} ({abm_code}): {len(abm_data)} records")
            
            # Calculate metrics for this ABM following the template structure
            
            # Basic counts - using correct columns
            unique_tbms = abm_data['TBM EMAIL_ID'].nunique() if 'TBM EMAIL_ID' in abm_data.columns else 0
            unique_hcps = abm_data['Doctor: Customer Code'].nunique()
            unique_requests = abm_data['Assigned Request Ids'].nunique()
            
            print(f"         Unique TBMs: {unique_tbms}, Unique HCPs: {unique_hcps}, Unique Requests: {unique_requests}")
            
            # HO Section (A + B) - Using Request Status field from Sample Master Tracker
            request_cancelled_out_of_stock = abm_data[abm_data['Request Status'].isin(['Out of stock', 'On hold', 'Not permitted'])]['Assigned Request Ids'].nunique()
            action_pending_at_ho = abm_data[abm_data['Request Status'].isin(['Request Raised'])]['Assigned Request Ids'].nunique()
            
            # HUB Section (D + E + F) - Using Request Status field
            # Pending for Invoicing (D) - Action Pending / in Process
            pending_for_invoicing = abm_data[abm_data['Request Status'].isin(['Action pending / In Process'])]['Assigned Request Ids'].nunique()
            # Pending for Dispatch (E) - Dispatch Pending
            pending_for_dispatch = abm_data[abm_data['Request Status'].isin(['Dispatch  Pending'])]['Assigned Request Ids'].nunique()
            
            # Delivery Status (G + H + I) - Using Request Status field
            delivered = abm_data[abm_data['Request Status'].isin(['Delivered'])]['Assigned Request Ids'].nunique()
            dispatched_in_transit = abm_data[abm_data['Request Status'].isin(['Dispatched & In Transit'])]['Assigned Request Ids'].nunique()
            rto = abm_data[abm_data['Request Status'].isin(['RTO'])]['Assigned Request Ids'].nunique()
            
            # RTO Reasons column mappings - using string contains for flexibility with hidden characters
            incomplete_address = abm_data[abm_data['Rto Reason'].str.contains('Incomplete Address', na=False)]['Assigned Request Ids'].nunique()
            doctor_non_contactable = abm_data[abm_data['Rto Reason'].str.contains('Dr. Non contactable', na=False)]['Assigned Request Ids'].nunique()
            doctor_refused_to_accept = abm_data[abm_data['Rto Reason'].str.contains('Doctor Refused to Accept', na=False)]['Assigned Request Ids'].nunique()
            
            # Calculated fields following the formulas from template
            requests_dispatched = delivered + dispatched_in_transit + rto  # F = G + H + I
            sent_to_hub = pending_for_invoicing + pending_for_dispatch + requests_dispatched  # C = D + E + F
            requests_raised = request_cancelled_out_of_stock + action_pending_at_ho + sent_to_hub  # A + B + C
            
            hold_delivery = 0
            
            # Create Area Name: "ABM Terr Code and ABM HQ" as per template
            # Use ABM HQ if available, otherwise use TBM HQ
            if 'ABM HQ' in abm_row and pd.notna(abm_row['ABM HQ']):
                abm_hq = abm_row['ABM HQ']
                print(f"      Using ABM HQ: {abm_hq} for {abm_name}")
            else:
                abm_hq = tbm_hq  # Fallback to TBM HQ
                print(f"      Using TBM HQ (fallback): {abm_hq} for {abm_name}")
            area_name = f"{abm_code} and {abm_hq}"
            
            summary_data.append({
                'Area Name': area_name,
                'ABM Name': abm_name,
                'Unique TBMs': unique_tbms,
                'Unique HCPs': unique_hcps,
                'Requests Raised': requests_raised,
                'Request Cancelled Out of Stock': request_cancelled_out_of_stock,
                'Action Pending at HO': action_pending_at_ho,
                'Sent to HUB': sent_to_hub,
                'Pending for Invoicing': pending_for_invoicing,
                'Pending for Dispatch': pending_for_dispatch,
                'Requests Dispatched': requests_dispatched,
                'Delivered': delivered,
                'Dispatched In Transit': dispatched_in_transit,
                'RTO': rto,
                'Incomplete Address': incomplete_address,
                'Doctor Non Contactable': doctor_non_contactable,
                'Doctor Refused to Accept': doctor_refused_to_accept,
                'Hold Delivery': hold_delivery
            })
            
            print(f"         Final counts - TBMs: {unique_tbms}, HCPs: {unique_hcps}, Requests: {unique_requests}, Raised: {requests_raised}")
        
        # Create DataFrame for this ZBM
        zbm_summary_df = pd.DataFrame(summary_data)
        
        # Create Excel file for this ZBM
        create_zbm_excel_report(zbm_code, zbm_name, zbm_email, zbm_summary_df, output_dir)
    
    print(f"\n🎉 Successfully created {len(zbms)} ZBM reports in directory: {output_dir}")

def create_zbm_excel_report(zbm_code, zbm_name, zbm_email, summary_df, output_dir):
    """Create Excel report for a specific ZBM with perfect formatting"""
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from copy import copy as copy_style

        # Load template
        wb = load_workbook('zbm_summary.xlsx')
        ws = wb['ZBM']

        print(f"   📋 Creating Excel report for {zbm_code}...")

        # Clear data area (rows 8 onwards) - preserve headers in row 7
        data_start_row = 7  # Headers are in row 7, data starts from row 8
        max_clear_rows = max(len(summary_df) + 10, 100)
        
        # Handle merged cells properly - remove ALL merged cells to avoid issues
        try:
            merged_ranges_to_remove = []
            for merged_range in ws.merged_cells.ranges:
                merged_ranges_to_remove.append(merged_range)
            
            # Remove all merged cells to avoid read-only issues
            for merged_range in merged_ranges_to_remove:
                try:
                    ws.unmerge_cells(str(merged_range))
                except Exception as e:
                    print(f"      Warning: Could not unmerge cell {merged_range}: {e}")
                    continue
        except Exception as e:
            print(f"      Warning: Error handling merged cells: {e}")
            # Continue anyway - we'll handle it in cell writing
        
        # Clear data area
        for r in range(data_start_row + 1, data_start_row + max_clear_rows):
            for c in range(5, 23):  # Columns E to V (matching template structure)
                cell = ws.cell(row=r, column=c)
                cell.value = None

        # Define exact column mapping based on template structure
        # Based on the actual template: Area Name (E), ABM Name (F), # Unique TBMs (G), etc.
        column_mapping = {
            'Area Name': 5,           # Column E - Area Name
            'ABM Name': 6,           # Column F - ABM Name  
            'Unique TBMs': 7,        # Column G - # Unique TBMs
            'Unique HCPs': 8,        # Column H - # Unique HCPs
            'Requests Raised': 9,     # Column I - # Requests Raised (A+B+C)
            'Request Cancelled Out of Stock': 10,  # Column J - Request Cancelled / Out of Stock (A)
            'Action Pending at HO': 11,            # Column K - Action pending / In Process At HO (B)
            'Sent to HUB': 12,                   # Column L - Sent to HUB (C)
            'Pending for Invoicing': 13,         # Column M - Pending for Invoicing (D)
            'Pending for Dispatch': 14,          # Column N - Pending for Dispatch (E)
            'Requests Dispatched': 15,           # Column O - # Requests Dispatched (F)
            'Delivered': 16,                     # Column P - Delivered (G)
            'Dispatched In Transit': 17,         # Column Q - Dispatched & In Transit (H)
            'RTO': 18,                           # Column R - RTO (I)
            'Incomplete Address': 19,            # Column S - Incomplete Address
            'Doctor Non Contactable': 20,        # Column T - Doctor Non Contactable
            'Doctor Refused to Accept': 21,      # Column U - Doctor Refused to Accept
            'Hold Delivery': 22                 # Column V - Hold Delivery
        }

        def copy_row_style(src_row_idx, dst_row_idx):
            """Copy formatting from source row to destination row"""
            for c in range(5, 23):  # Columns E to V (matching template structure)
                src = ws.cell(row=src_row_idx, column=c)
                dst = ws.cell(row=dst_row_idx, column=c)
                
                if src.font:
                    dst.font = copy_style(src.font)
                if src.alignment:
                    dst.alignment = copy_style(src.alignment)
                if src.border:
                    dst.border = copy_style(src.border)
                if src.fill:
                    dst.fill = copy_style(src.fill)
                dst.number_format = src.number_format

        def write_to_cell_safely(row, col, value, formatting_func=None):
            """Write to a cell safely"""
            try:
                cell = ws.cell(row=row, column=col)
                # Check if cell is merged and handle accordingly
                if hasattr(cell, 'value') and hasattr(cell.value, '__class__') and 'MergedCell' in str(cell.value.__class__):
                    # Skip merged cells
                    return None
                cell.value = value
                
                if formatting_func:
                    formatting_func(cell)
                
                return cell
            except Exception as e:
                print(f"      Warning: Could not write to cell ({row}, {col}): {e}")
                return None

        # Write data rows
        for i in range(len(summary_df)):
            target_row = data_start_row + 1 + i  # Start from row 8
            if target_row > ws.max_row:
                ws.insert_rows(target_row)
            
            # Copy formatting from template row 8 (first data row)
            copy_row_style(8, target_row)
            
            # Write data according to exact column mapping
            for col_name, col_num in column_mapping.items():
                if col_name in summary_df.columns:
                    value = summary_df.at[i, col_name]
                    
                    def apply_number_formatting(cell):
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            cell.number_format = '0'  # Integer format
                    
                    write_to_cell_safely(target_row, col_num, value, apply_number_formatting)

        # Add total row
        total_row = data_start_row + 1 + len(summary_df)
        if total_row > ws.max_row:
            ws.insert_rows(total_row)
        
        # Copy formatting for total row
        copy_row_style(8, total_row)
        
        # Write totals
        def apply_total_formatting(cell):
            cell.font = Font(bold=True, name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        write_to_cell_safely(total_row, 5, None)  # Empty first column
        write_to_cell_safely(total_row, 6, "Total", apply_total_formatting)
        
        # Calculate and write totals for each column
        for col_name, col_num in column_mapping.items():
            if col_name in summary_df.columns and col_name not in ['Area Name', 'ABM Name']:
                total_value = summary_df[col_name].sum()
                
                def apply_total_value_formatting(cell):
                    cell.font = Font(bold=True, name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if isinstance(total_value, (int, float)) and not pd.isna(total_value):
                        cell.number_format = '0'  # Integer format
                
                write_to_cell_safely(total_row, col_num, total_value, apply_total_value_formatting)
        
        # Add merged cells for better formatting (matching template structure)
        try:
            # Merge cells in header row for better visual grouping
            # Merge cells for "Requests Raised (A+B+C)" - columns I
            ws.merge_cells(f'I7:I7')
            
            # Merge cells for "Request Cancelled / Out of Stock (A)" - columns J
            ws.merge_cells(f'J7:J7')
            
            # Merge cells for "Action pending / In Process At HO (B)" - columns K
            ws.merge_cells(f'K7:K7')
            
            # Merge cells for "Sent to HUB (C) (D+E+F)" - columns L
            ws.merge_cells(f'L7:L7')
            
            # Merge cells for "Pending for Invoicing (D)" - columns M
            ws.merge_cells(f'M7:M7')
            
            # Merge cells for "Pending for Dispatch (E)" - columns N
            ws.merge_cells(f'N7:N7')
            
            # Merge cells for "# Requests Dispatched (F) (G+H+I)" - columns O
            ws.merge_cells(f'O7:O7')
            
            # Merge cells for "Delivered (G)" - columns P
            ws.merge_cells(f'P7:P7')
            
            # Merge cells for "Dispatched & In Transit (H)" - columns Q
            ws.merge_cells(f'Q7:Q7')
            
            # Merge cells for "RTO (I)" - columns R
            ws.merge_cells(f'R7:R7')
            
            print(f"   ✅ Applied merged cell formatting")
            
        except Exception as e:
            print(f"   ⚠️ Could not apply merged cell formatting: {e}")

        # Save file - handle special characters in ZBM name
        safe_zbm_name = str(zbm_name).replace(' ', '_').replace('/', '_').replace('\\', '_')
        filename = f"ZBM_Summary_{zbm_code}_{safe_zbm_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        wb.save(filepath)
        print(f"   ✅ Created: {filename}")
        
        # Print summary statistics
        print(f"   📊 Summary for {zbm_code}:")
        print(f"      Total ABMs: {len(summary_df)}")
        print(f"      Total Unique TBMs: {summary_df['Unique TBMs'].sum()}")
        print(f"      Total Unique HCPs: {summary_df['Unique HCPs'].sum()}")
        print(f"      Total Requests Raised: {summary_df['Requests Raised'].sum()}")
        print(f"      Total Delivered: {summary_df['Delivered'].sum()}")
        print(f"      Total RTO: {summary_df['RTO'].sum()}")
        
    except Exception as e:
        print(f"   ❌ Error creating Excel report for {zbm_code}: {e}")

if __name__ == "__main__":
    create_zbm_hierarchical_reports()
