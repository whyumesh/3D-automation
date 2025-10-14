import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_corrected_zbm_reports():
    """
    Create ZBM reports with EXACT template format and CORRECT RTO data mapping
    """
    
    print("🔄 Starting Corrected ZBM Reports Generation...")
    
    # Read master tracker data
    print("📖 Reading master_tracker.csv...")
    try:
        df = pd.read_csv('master_tracker.csv', encoding='latin-1', low_memory=False)
        print(f"✅ Successfully loaded {len(df)} records from master_tracker.csv")
    except Exception as e:
        print(f"❌ Error reading master_tracker.csv: {e}")
        return
    
    # Clean and prepare data
    print("🧹 Cleaning and preparing data...")
    
    # Ensure required columns exist
    required_columns = ['ZBM Terr Code', 'ZBM Name', 'ZBM EMAIL_ID',
                        'ABM Terr Code', 'ABM Name', 'ABM EMAIL_ID',
                        'TBM HQ', 'TBM EMAIL_ID',
                        'Doctor: Customer Code', 'Assigned Request Ids', 'Request Status', 'Rto Reason']
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        print(f"❌ Missing required columns in master_tracker.csv: {missing}")
        return

    # Remove rows where key fields are null or empty
    df = df.dropna(subset=['ZBM Terr Code', 'ZBM Name', 'ABM Terr Code', 'ABM Name', 'TBM HQ'])
    df = df[df['ZBM Terr Code'].astype(str).str.strip() != '']
    df = df[df['ABM Terr Code'].astype(str).str.strip() != '']
    df = df[df['TBM HQ'].astype(str).str.strip() != '']

    print(f"📊 After cleaning: {len(df)} records remaining")

    # Compute Final Answer per unique request id using corrected rules
    print("🧠 Computing final status per unique Request Id using corrected rules...")
    try:
        def normalize(text):
            normalized = str(text).strip().casefold()
            normalized = normalized.replace('  ', ' ')  # Fix spacing issues
            return normalized

        # Load rules from logic.xlsx
        xls_rules = pd.ExcelFile('logic.xlsx')
        sheet2 = pd.read_excel(xls_rules, 'Sheet2')

        rules = {}
        for _, row in sheet2.iterrows():
            statuses = [normalize(s) for s in row.drop('Final Answer').dropna().tolist()]
            statuses = tuple(sorted(set(statuses)))
            rules[statuses] = row['Final Answer']

        # Add missing rules identified from validation
        additional_rules = {
            ('action pending / in process', 'delivered'): 'Delivered',
            ('action pending / in process', 'dispatched & in transit'): 'Dispatched & In Transit',
            ('action pending / in process', 'dispatch pending'): 'Dispatch Pending',
            ('action pending / in process', 'out of stock'): 'Out of stock',
            ('action pending / in process', 'return'): 'Return',
            ('delivered', 'return'): 'Delivered',
            ('dispatch pending', 'delivered'): 'Delivered',
            ('dispatch pending', 'dispatched & in transit'): 'Dispatched & In Transit',
            ('dispatch pending', 'return'): 'Return',
            ('dispatched & in transit', 'return'): 'Dispatched & In Transit',
            ('out of stock', 'return'): 'Out of stock',
            ('request raised', 'action pending / in process'): 'Action pending / In Process',
            ('request raised', 'delivered'): 'Delivered',
            ('request raised', 'dispatch pending'): 'Dispatch Pending',
            ('request raised', 'dispatched & in transit'): 'Dispatched & In Transit',
            ('request raised', 'out of stock'): 'Out of stock',
            ('request raised', 'return'): 'Return',
        }
        
        rules.update(additional_rules)

        # Apply rules to compute Final Answer
        def compute_final_answer(request_id):
            request_data = df[df['Assigned Request Ids'] == request_id]
            if len(request_data) == 0:
                return 'Unknown'
            
            statuses = [normalize(s) for s in request_data['Request Status'].dropna().unique()]
            statuses = tuple(sorted(set(statuses)))
            
            if statuses in rules:
                return rules[statuses]
            else:
                # If no rule found, return the most common status
                return request_data['Request Status'].mode().iloc[0] if len(request_data['Request Status'].mode()) > 0 else 'Unknown'

        # Apply final answer computation
        unique_requests = df['Assigned Request Ids'].unique()
        print(f"🔍 Computing final answers for {len(unique_requests)} unique requests...")
        
        final_answers = {}
        for req_id in unique_requests:
            final_answers[req_id] = compute_final_answer(req_id)
        
        # Map final answers back to dataframe
        df['Final Answer'] = df['Assigned Request Ids'].map(final_answers)
        
        print("✅ Final Answer computation completed")
        
    except Exception as e:
        print(f"❌ Error computing final answers: {e}")
        return

    # Get unique ZBMs
    zbms = df[['ZBM Terr Code', 'ZBM Name', 'ZBM EMAIL_ID']].drop_duplicates().sort_values('ZBM Terr Code')
    print(f"📋 Found {len(zbms)} unique ZBMs")

    # Create output directory
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = f"Corrected_ZBM_Reports_{timestamp}"
    import os
    os.makedirs(output_dir, exist_ok=True)
    
    successful_reports = 0
    failed_reports = 0
    
    # Process each ZBM
    for i, (_, zbm_row) in enumerate(zbms.iterrows()):
        zbm_code = zbm_row['ZBM Terr Code']
        zbm_name = zbm_row['ZBM Name']
        zbm_email = zbm_row['ZBM EMAIL_ID']
        
        # Handle NaN names
        if pd.isna(zbm_name):
            zbm_name = "Unknown"
        
        print(f"📋 Processing ZBM {i+1}/{len(zbms)}: {zbm_name} ({zbm_code})")
        
        try:
            # Filter data for this ZBM
            zbm_data = df[df['ZBM Terr Code'] == zbm_code].copy()
            
            if len(zbm_data) == 0:
                print(f"⚠️ No data found for ZBM: {zbm_code}")
                failed_reports += 1
                continue
            
            # Get unique ABMs under this ZBM
            abms = zbm_data[['ABM Terr Code', 'ABM Name', 'ABM EMAIL_ID']].drop_duplicates().sort_values('ABM Terr Code')
            
            # Create summary data following the EXACT template structure
            summary_data = []
            
            for _, abm_row in abms.iterrows():
                abm_code = abm_row['ABM Terr Code']
                abm_name = abm_row['ABM Name']
                abm_email = abm_row['ABM EMAIL_ID']
                
                abm_data = zbm_data[zbm_data['ABM Terr Code'] == abm_code]
                
                # Calculate metrics for this ABM following the EXACT template structure
                
                # Basic counts
                unique_tbms = abm_data['TBM EMAIL_ID'].nunique()
                unique_hcps = abm_data['Doctor: Customer Code'].nunique()
                
                # HO Section (A + B)
                request_cancelled_out_of_stock = abm_data[abm_data['Final Answer'].isin(['Out of stock', 'On hold', 'Not permitted'])]['Assigned Request Ids'].nunique()
                action_pending_at_ho = abm_data[abm_data['Final Answer'].isin(['Action pending / In Process'])]['Assigned Request Ids'].nunique()
                
                # HUB Section (D + E + F)
                pending_for_invoicing = 0  # Placeholder - would need specific logic
                pending_for_dispatch = abm_data[abm_data['Final Answer'].isin(['Dispatch Pending'])]['Assigned Request Ids'].nunique()
                
                # Delivery Status (G + H + I)
                delivered = abm_data[abm_data['Final Answer'].isin(['Delivered'])]['Assigned Request Ids'].nunique()
                dispatched_in_transit = abm_data[abm_data['Final Answer'].isin(['Dispatched & In Transit'])]['Assigned Request Ids'].nunique()
                
                # CORRECTED RTO CALCULATION - Use actual RTO data
                rto = abm_data[abm_data['Rto Reason'].notna()]['Assigned Request Ids'].nunique()
                
                # Calculated fields following the formulas from template
                requests_dispatched = delivered + dispatched_in_transit + rto  # F = G + H + I
                sent_to_hub = pending_for_invoicing + pending_for_dispatch + requests_dispatched  # C = D + E + F
                requests_raised = request_cancelled_out_of_stock + action_pending_at_ho + sent_to_hub  # A + B + C
                
                # CORRECTED RTO Reasons calculations using actual RTO data
                rto_data = abm_data[abm_data['Rto Reason'].notna()]
                
                # Map RTO reasons to template categories
                incomplete_address = rto_data[rto_data['Rto Reason'].str.contains('Incomplete Address', case=False, na=False)]['Assigned Request Ids'].nunique()
                doctor_non_contactable = rto_data[rto_data['Rto Reason'].str.contains('Non contactable', case=False, na=False)]['Assigned Request Ids'].nunique()
                doctor_refused_to_accept = rto_data[rto_data['Rto Reason'].str.contains('refused to accept', case=False, na=False)]['Assigned Request Ids'].nunique()
                hold_delivery = rto_data[rto_data['Rto Reason'].str.contains('Hold Delivery', case=False, na=False)]['Assigned Request Ids'].nunique()
                
                # Create Area Name: "ABM Terr Code - ABM Name"
                area_name = f"{abm_code} - {abm_name}"
                
                summary_row = {
                    'Area Name': area_name,
                    'ABM Name': abm_name,
                    '# Unique TBMs': unique_tbms,
                    '# Unique HCPs': unique_hcps,
                    '# Requests Raised (A+B+C)': requests_raised,
                    'Request Cancelled / Out of Stock (A)': request_cancelled_out_of_stock,
                    'Action pending / In Process At HO (B)': action_pending_at_ho,
                    "Sent to HUB ('C) (D+E+F)": sent_to_hub,
                    'Pending for Invoicing (D)': pending_for_invoicing,
                    'Pending for Dispatch (E)': pending_for_dispatch,
                    '# Requests Dispatched (F) (G+H+I)': requests_dispatched,
                    'Delivered (G)': delivered,
                    'Dispatched & In Transit (H)': dispatched_in_transit,
                    'RTO (I)': rto,
                    'Incomplete Address': incomplete_address,
                    'Doctor Non Contactable': doctor_non_contactable,
                    'Doctor Refused to Accept': doctor_refused_to_accept,
                    'Hold Delivery': hold_delivery
                }
                
                summary_data.append(summary_row)
            
            # Create Excel file with EXACT template format
            wb = Workbook()
            ws = wb.active
            ws.title = "ZBM Summary"
            
            # Set EXACT template headers (Row 0)
            headers_row0 = [
                None, 'Area Name', 'ABM Name', '# Unique TBMs', '# Unique HCPs', 
                '# Requests Raised\n(A+B+C)', 'HO', None, None, 'HUB', None, None, 
                'Delivery Status', None, None, 'RTO Reasons', None, None, None
            ]
            
            for col, header in enumerate(headers_row0, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Set EXACT template headers (Row 1)
            headers_row1 = [
                None, None, None, None, None, None,
                'Request Cancelled / Out of Stock (A)', 'Action pending / In Process At HO (B)', 
                "Sent to HUB ('C)\n(D+E+F)", 'Pending for Invoicing (D)', 'Pending for Dispatch (E)', 
                '# Requests Dispatched (F)\n(G+H+I)', 'Delivered (G)', 'Dispatched & In Transit (H)', 
                'RTO (I)', 'Incomplete Address', 'Doctor Non Contactable', 'Doctor Refused to Accept', 'Hold Delivery'
            ]
            
            for col, header in enumerate(headers_row1, 1):
                ws.cell(row=2, column=col, value=header)
            
            # Add data rows starting from row 3
            for row_idx, data in enumerate(summary_data, 3):
                ws.cell(row=row_idx, column=2, value=data['Area Name'])
                ws.cell(row=row_idx, column=3, value=data['ABM Name'])
                ws.cell(row=row_idx, column=4, value=data['# Unique TBMs'])
                ws.cell(row=row_idx, column=5, value=data['# Unique HCPs'])
                ws.cell(row=row_idx, column=6, value=data['# Requests Raised (A+B+C)'])
                ws.cell(row=row_idx, column=7, value=data['Request Cancelled / Out of Stock (A)'])
                ws.cell(row=row_idx, column=8, value=data['Action pending / In Process At HO (B)'])
                ws.cell(row=row_idx, column=9, value=data["Sent to HUB ('C) (D+E+F)"])
                ws.cell(row=row_idx, column=10, value=data['Pending for Invoicing (D)'])
                ws.cell(row=row_idx, column=11, value=data['Pending for Dispatch (E)'])
                ws.cell(row=row_idx, column=12, value=data['# Requests Dispatched (F) (G+H+I)'])
                ws.cell(row=row_idx, column=13, value=data['Delivered (G)'])
                ws.cell(row=row_idx, column=14, value=data['Dispatched & In Transit (H)'])
                ws.cell(row=row_idx, column=15, value=data['RTO (I)'])
                ws.cell(row=row_idx, column=16, value=data['Incomplete Address'])
                ws.cell(row=row_idx, column=17, value=data['Doctor Non Contactable'])
                ws.cell(row=row_idx, column=18, value=data['Doctor Refused to Accept'])
                ws.cell(row=row_idx, column=19, value=data['Hold Delivery'])
            
            # Add totals row
            if summary_data:
                totals_row = len(summary_data) + 3
                ws.cell(row=totals_row, column=2, value="TOTAL")
                ws.cell(row=totals_row, column=4, value=sum(d['# Unique TBMs'] for d in summary_data))
                ws.cell(row=totals_row, column=5, value=sum(d['# Unique HCPs'] for d in summary_data))
                ws.cell(row=totals_row, column=6, value=sum(d['# Requests Raised (A+B+C)'] for d in summary_data))
                ws.cell(row=totals_row, column=7, value=sum(d['Request Cancelled / Out of Stock (A)'] for d in summary_data))
                ws.cell(row=totals_row, column=8, value=sum(d['Action pending / In Process At HO (B)'] for d in summary_data))
                ws.cell(row=totals_row, column=9, value=sum(d["Sent to HUB ('C) (D+E+F)"] for d in summary_data))
                ws.cell(row=totals_row, column=10, value=sum(d['Pending for Invoicing (D)'] for d in summary_data))
                ws.cell(row=totals_row, column=11, value=sum(d['Pending for Dispatch (E)'] for d in summary_data))
                ws.cell(row=totals_row, column=12, value=sum(d['# Requests Dispatched (F) (G+H+I)'] for d in summary_data))
                ws.cell(row=totals_row, column=13, value=sum(d['Delivered (G)'] for d in summary_data))
                ws.cell(row=totals_row, column=14, value=sum(d['Dispatched & In Transit (H)'] for d in summary_data))
                ws.cell(row=totals_row, column=15, value=sum(d['RTO (I)'] for d in summary_data))
                ws.cell(row=totals_row, column=16, value=sum(d['Incomplete Address'] for d in summary_data))
                ws.cell(row=totals_row, column=17, value=sum(d['Doctor Non Contactable'] for d in summary_data))
                ws.cell(row=totals_row, column=18, value=sum(d['Doctor Refused to Accept'] for d in summary_data))
                ws.cell(row=totals_row, column=19, value=sum(d['Hold Delivery'] for d in summary_data))
            
            # Apply formatting
            def apply_formatting():
                # Header formatting
                header_font = Font(bold=True, size=11)
                header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Apply header formatting to rows 1 and 2
                for row in [1, 2]:
                    for col in range(1, 20):
                        cell = ws.cell(row=row, column=col)
                        if cell.value is not None:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                
                # Data formatting
                data_alignment = Alignment(horizontal="center", vertical="center")
                for row in range(3, totals_row + 1):
                    for col in range(2, 20):
                        cell = ws.cell(row=row, column=col)
                        if cell.value is not None:
                            cell.alignment = data_alignment
                
                # Auto-adjust column widths
                for col in range(1, 20):
                    max_length = 0
                    column = ws.column_dimensions[chr(64 + col)]
                    for row in range(1, totals_row + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    column.width = min(max_length + 2, 50)
            
            apply_formatting()
            
            # Save file
            filename = f"ZBM_Summary_{zbm_code}_{zbm_name.replace(' ', '_')}_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            wb.save(filepath)
            
            print(f"✅ Created report: {filename}")
            print(f"   📊 ABMs: {len(summary_data)}, Total RTO: {sum(d['RTO (I)'] for d in summary_data)}")
            print(f"   📈 RTO Reasons - Incomplete: {sum(d['Incomplete Address'] for d in summary_data)}, Non-contactable: {sum(d['Doctor Non Contactable'] for d in summary_data)}, Refused: {sum(d['Doctor Refused to Accept'] for d in summary_data)}")
            
            successful_reports += 1
            
        except Exception as e:
            print(f"❌ Error processing ZBM {zbm_code}: {e}")
            failed_reports += 1
            continue
    
    print(f"\n🎉 ZBM Reports Generation Completed!")
    print(f"✅ Successful: {successful_reports}")
    print(f"❌ Failed: {failed_reports}")
    print(f"📁 Output Directory: {output_dir}")
    
    return output_dir

if __name__ == "__main__":
    create_corrected_zbm_reports()
