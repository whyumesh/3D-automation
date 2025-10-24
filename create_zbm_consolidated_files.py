import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy as copy_style
import warnings

warnings.filterwarnings('ignore', category=FutureWarning, module='pandas')

def create_zbm_hierarchical_reports():
    """
    Create separate ZBM reports showing ABM hierarchy with perfect tallies
    FIXED VERSION with diagnostic output
    """
    
    print("🔄 Starting ZBM Hierarchical Reports Creation (FIXED VERSION)...")
    
    # Read master tracker data
    print("📖 Reading ZBM Automation Email 2410252.xlsx...")
    try:
        df = pd.read_excel('ZBM Automation Email 2410252.xlsx')
        print(f"✅ Successfully loaded {len(df)} records")
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        return
    
    # Ensure required columns exist
    required_columns = ['ZBM Terr Code', 'ZBM Name', 'ZBM EMAIL_ID',
                        'ABM Terr Code', 'ABM Name', 'ABM EMAIL_ID',
                        'TBM HQ', 'TBM EMAIL_ID',
                        'Doctor: Customer Code', 'Assigned Request Ids', 'Request Status', 'Rto Reason']
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        print(f"❌ Missing required columns: {missing}")
        return

    # Clean data
    print("🧹 Cleaning data...")
    df = df.dropna(subset=['ZBM Terr Code', 'ZBM Name', 'ABM Terr Code', 'ABM Name', 'TBM HQ'])
    df = df[df['ZBM Terr Code'].astype(str).str.strip() != '']
    df = df[df['ABM Terr Code'].astype(str).str.strip() != '']
    df = df[df['TBM HQ'].astype(str).str.strip() != '']
    print(f"📊 After cleaning: {len(df)} records")

    # ===== CRITICAL FIX: Compute Final Answer with better debugging =====
    print("\n🧠 Computing Final Answer per Request ID...")
    try:
        xls_rules = pd.ExcelFile('logic.xlsx')
        sheet2 = pd.read_excel(xls_rules, 'Sheet2')

        def normalize(text):
            return str(text).strip().casefold()

        # Build rules dictionary
        rules = {}
        for _, row in sheet2.iterrows():
            statuses = [normalize(s) for s in row.drop('Final Answer').dropna().tolist()]
            statuses = tuple(sorted(set(statuses)))
            rules[statuses] = row['Final Answer']
        
        print(f"   Loaded {len(rules)} rules from logic.xlsx")

        # Group by request ID and get unique statuses
        grouped = df.groupby('Assigned Request Ids')['Request Status'].apply(
            lambda x: list(x.unique())
        ).reset_index()
        
        def get_final_answer(status_list):
            normalized = [normalize(s) for s in status_list]
            key = tuple(sorted(set(normalized)))
            result = rules.get(key, '❌ No matching rule')
            return result
        
        grouped['Final Answer'] = grouped['Request Status'].apply(get_final_answer)
        
        # Diagnostic: Show sample mappings
        print("\n   📋 Sample Request ID → Final Answer mappings:")
        for i in range(min(5, len(grouped))):
            req_id = grouped.iloc[i]['Assigned Request Ids']
            statuses = grouped.iloc[i]['Request Status']
            final = grouped.iloc[i]['Final Answer']
            print(f"      {req_id}: {statuses} → {final}")
        
        # Check for unmapped rules
        no_match = grouped[grouped['Final Answer'] == '❌ No matching rule']
        if len(no_match) > 0:
            print(f"\n   ⚠️ WARNING: {len(no_match)} request IDs have no matching rule!")
            print("   First few examples:")
            for i in range(min(3, len(no_match))):
                print(f"      {no_match.iloc[i]['Assigned Request Ids']}: {no_match.iloc[i]['Request Status']}")

        # Merge back to main dataframe
        # IMPORTANT: Keep only Request ID and Final Answer to avoid duplication
        df = df.merge(
            grouped[['Assigned Request Ids', 'Final Answer']], 
            on='Assigned Request Ids', 
            how='left'
        )
        
        print(f"   ✅ Final Answer computed for {df['Assigned Request Ids'].nunique()} unique requests")
        
    except Exception as e:
        print(f"❌ Error computing Final Answer: {e}")
        import traceback
        traceback.print_exc()
        return

    # Get unique ZBMs
    zbms = df[['ZBM Terr Code', 'ZBM Name', 'ZBM EMAIL_ID']].drop_duplicates().sort_values('ZBM Terr Code')
    print(f"\n📋 Found {len(zbms)} unique ZBMs")
    
    # Create output directory
    timestamp = datetime.now().strftime('%Y%m%d')
    output_dir = f"ZBM_Reports_{timestamp}_FIXED"
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Created output directory: {output_dir}")
    
    # Process each ZBM
    for idx, zbm_row in zbms.iterrows():
        zbm_code = zbm_row['ZBM Terr Code']
        zbm_name = zbm_row['ZBM Name']
        zbm_email = zbm_row['ZBM EMAIL_ID']
        
        print(f"\n{'='*60}")
        print(f"🔄 Processing ZBM: {zbm_code} - {zbm_name}")
        print(f"{'='*60}")
        
        # Filter data for this ZBM
        zbm_data = df[df['ZBM Terr Code'] == zbm_code].copy()
        
        if len(zbm_data) == 0:
            print(f"⚠️ No data found for ZBM: {zbm_code}")
            continue
        
        # Get unique ABMs under this ZBM
        abms = zbm_data[['ABM Terr Code', 'ABM Name', 'ABM EMAIL_ID']].drop_duplicates().sort_values('ABM Terr Code')
        print(f"   📊 Found {len(abms)} ABMs under this ZBM")
        
        # Create summary data for this ZBM
        summary_data = []
        
        # Track totals for verification
        zbm_totals = {
            'unique_requests': zbm_data['Assigned Request Ids'].nunique(),
            'unique_hcps': zbm_data['Doctor: Customer Code'].nunique(),
            'unique_tbms': zbm_data['TBM EMAIL_ID'].nunique() if 'TBM EMAIL_ID' in zbm_data.columns else 0
        }
        
        print(f"\n   🎯 ZBM-Level Totals (for verification):")
        print(f"      Unique Requests: {zbm_totals['unique_requests']}")
        print(f"      Unique HCPs: {zbm_totals['unique_hcps']}")
        print(f"      Unique TBMs: {zbm_totals['unique_tbms']}")
        
        for _, abm_row in abms.iterrows():
            abm_code = abm_row['ABM Terr Code']
            abm_name = abm_row['ABM Name']
            abm_email = abm_row['ABM EMAIL_ID']
            
            # Filter data for this specific ABM
            abm_data = zbm_data[
                (zbm_data['ABM Terr Code'] == abm_code) & 
                (zbm_data['ABM Name'] == abm_name)
            ].copy()
            
            if len(abm_data) == 0:
                continue
            
            print(f"\n   📍 Processing ABM: {abm_name} ({abm_code})")
            print(f"      Total rows: {len(abm_data)}")
            
            # Basic metrics
            unique_tbms = abm_data['TBM EMAIL_ID'].nunique() if 'TBM EMAIL_ID' in abm_data.columns else 0
            unique_hcps = abm_data['Doctor: Customer Code'].nunique()
            unique_requests = abm_data['Assigned Request Ids'].nunique()
            
            print(f"      Unique Requests: {unique_requests}")
            print(f"      Unique HCPs: {unique_hcps}")
            print(f"      Unique TBMs: {unique_tbms}")
            
            # ===== CRITICAL FIX: Count based on UNIQUE Request IDs only =====
            # Get unique request IDs with their Final Answer
            unique_req_data = abm_data[['Assigned Request Ids', 'Final Answer', 'Rto Reason']].drop_duplicates(subset=['Assigned Request Ids'])
            
            print(f"      Unique request records: {len(unique_req_data)}")
            
            # HO Section (A + B)
            request_cancelled_out_of_stock = unique_req_data[
                unique_req_data['Final Answer'].isin(['Out of stock', 'On hold', 'Not permitted'])
            ].shape[0]
            
            action_pending_at_ho = unique_req_data[
                unique_req_data['Final Answer'].isin(['Request Raised', 'Action pending / In Process At HO'])
            ].shape[0]
            
            # HUB Section (D + E)
            pending_for_invoicing = unique_req_data[
                unique_req_data['Final Answer'].isin(['Action pending / In Process At Hub'])
            ].shape[0]
            
            pending_for_dispatch = unique_req_data[
                unique_req_data['Final Answer'].isin(['Dispatch  Pending'])
            ].shape[0]
            
            # Delivery Status (G + H)
            delivered = unique_req_data[
                unique_req_data['Final Answer'].isin(['Delivered'])
            ].shape[0]
            
            dispatched_in_transit = unique_req_data[
                unique_req_data['Final Answer'].isin(['Dispatched & In Transit'])
            ].shape[0]
            
            # RTO Reasons (based on Rto Reason field)
            incomplete_address = unique_req_data[
                unique_req_data['Rto Reason'].str.contains('Incomplete Address', na=False, case=False)
            ].shape[0]
            
            doctor_non_contactable = unique_req_data[
                unique_req_data['Rto Reason'].str.contains('Dr. Non contactable', na=False, case=False)
            ].shape[0]
            
            doctor_refused_to_accept = unique_req_data[
                unique_req_data['Rto Reason'].str.contains('Doctor Refused to Accept', na=False, case=False)
            ].shape[0]
            
            rto_total = incomplete_address + doctor_non_contactable + doctor_refused_to_accept
            
            # Calculated fields
            requests_dispatched = delivered + dispatched_in_transit + rto_total
            sent_to_hub = pending_for_invoicing + pending_for_dispatch + requests_dispatched
            requests_raised = request_cancelled_out_of_stock + action_pending_at_ho + sent_to_hub
            
            # Diagnostic output
            print(f"      Status breakdown:")
            print(f"         A (Cancelled/Stock): {request_cancelled_out_of_stock}")
            print(f"         B (Pending HO): {action_pending_at_ho}")
            print(f"         D (Pending Invoice): {pending_for_invoicing}")
            print(f"         E (Pending Dispatch): {pending_for_dispatch}")
            print(f"         G (Delivered): {delivered}")
            print(f"         H (In Transit): {dispatched_in_transit}")
            print(f"         I (RTO): {rto_total}")
            print(f"      Calculated:")
            print(f"         F (Dispatched): {requests_dispatched} = {delivered}+{dispatched_in_transit}+{rto_total}")
            print(f"         C (Sent to Hub): {sent_to_hub} = {pending_for_invoicing}+{pending_for_dispatch}+{requests_dispatched}")
            print(f"         Total Raised: {requests_raised} = {request_cancelled_out_of_stock}+{action_pending_at_ho}+{sent_to_hub}")
            print(f"      Verification: Total Raised ({requests_raised}) should equal Unique Requests ({unique_requests})")
            
            if requests_raised != unique_requests:
                print(f"      ⚠️ MISMATCH DETECTED! Difference: {requests_raised - unique_requests}")
                # Show Final Answer distribution
                print(f"      Final Answer distribution:")
                for fa in unique_req_data['Final Answer'].unique():
                    count = (unique_req_data['Final Answer'] == fa).sum()
                    print(f"         {fa}: {count}")
            
            # Get ABM HQ or use TBM HQ as fallback
            tbm_hq = abm_data['TBM HQ'].iloc[0] if len(abm_data) > 0 else ''
            abm_hq = abm_data['ABM HQ'].iloc[0] if 'ABM HQ' in abm_data.columns and len(abm_data) > 0 else tbm_hq
            area_name = f"{abm_code} - {abm_hq}"
            
            summary_data.append({
                'Area Name': area_name,
                'ABM Name': abm_name,
                'Unique TBMs': unique_tbms,
                'Unique HCPs': unique_hcps,
                'Requests Raised': requests_raised,
                'Request Cancelled Out of Stock': request_cancelled_out_of_stock,
                'Action Pending at HO': action_pending_at_ho,
                'Sent to HUB': sent_to_hub,
                'Pending for Invoicing': pending_for_invoicing,
                'Pending for Dispatch': pending_for_dispatch,
                'Requests Dispatched': requests_dispatched,
                'Delivered': delivered,
                'Dispatched In Transit': dispatched_in_transit,
                'RTO': rto_total,
                'Incomplete Address': incomplete_address,
                'Doctor Non Contactable': doctor_non_contactable,
                'Doctor Refused to Accept': doctor_refused_to_accept,
                'Hold Delivery': 0
            })
        
        # Create DataFrame
        zbm_summary_df = pd.DataFrame(summary_data)
        
        # Verify totals
        print(f"\n   ✅ Summary verification for ZBM {zbm_code}:")
        print(f"      Sum of Requests Raised: {zbm_summary_df['Requests Raised'].sum()}")
        print(f"      Expected (Unique Requests): {zbm_totals['unique_requests']}")
        if zbm_summary_df['Requests Raised'].sum() != zbm_totals['unique_requests']:
            print(f"      ⚠️ TOTAL MISMATCH! Difference: {zbm_summary_df['Requests Raised'].sum() - zbm_totals['unique_requests']}")
        
        # Create Excel report
        create_zbm_excel_report(zbm_code, zbm_name, zbm_email, zbm_summary_df, output_dir)
    
    print(f"\n🎉 Successfully created {len(zbms)} ZBM reports in: {output_dir}")

def create_zbm_excel_report(zbm_code, zbm_name, zbm_email, summary_df, output_dir):
    """Create Excel report for a specific ZBM"""
    
    try:
        wb = load_workbook('zbm_summary.xlsx')
        ws = wb['ZBM']

        def get_cell_value_handling_merged(row, col):
            cell = ws.cell(row=row, column=col)
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    return top_left_cell.value
            return cell.value
        
        # Find header row
        header_row = None
        for row_idx in range(1, 15):
            for col_idx in range(1, 30):
                cell_value = get_cell_value_handling_merged(row_idx, col_idx)
                if cell_value and 'Area Name' in str(cell_value):
                    header_row = row_idx
                    break
            if header_row:
                break
        
        if header_row is None:
            header_row = 7
        
        data_start_row = header_row + 1
        
        # Map columns
        column_mapping = {}
        for col_idx in range(1, 30):
            header_val = get_cell_value_handling_merged(header_row, col_idx)
            if header_val:
                header_str = str(header_val).strip()
                
                if 'Area Name' in header_str:
                    column_mapping['Area Name'] = col_idx
                elif 'ABM Name' in header_str:
                    column_mapping['ABM Name'] = col_idx
                elif 'Unique TBMs' in header_str or '# Unique TBMs' in header_str:
                    column_mapping['Unique TBMs'] = col_idx
                elif 'Unique HCPs' in header_str or '# Unique HCPs' in header_str:
                    column_mapping['Unique HCPs'] = col_idx
                elif 'Requests Raised' in header_str or '# Requests Raised' in header_str:
                    column_mapping['Requests Raised'] = col_idx
                elif 'Request Cancelled' in header_str or 'Out of Stock' in header_str:
                    column_mapping['Request Cancelled Out of Stock'] = col_idx
                elif 'Action pending' in header_str and 'HO' in header_str:
                    column_mapping['Action Pending at HO'] = col_idx
                elif 'Sent to HUB' in header_str:
                    column_mapping['Sent to HUB'] = col_idx
                elif 'Pending for Invoicing' in header_str:
                    column_mapping['Pending for Invoicing'] = col_idx
                elif 'Pending for Dispatch' in header_str:
                    column_mapping['Pending for Dispatch'] = col_idx
                elif 'Requests Dispatched' in header_str or '# Requests Dispatched' in header_str:
                    column_mapping['Requests Dispatched'] = col_idx
                elif header_str == 'Delivered' or 'Delivered (G)' in header_str:
                    column_mapping['Delivered'] = col_idx
                elif 'Dispatched & In Transit' in header_str or 'Dispatched In Transit' in header_str:
                    column_mapping['Dispatched In Transit'] = col_idx
                elif header_str == 'RTO' or 'RTO (I)' in header_str:
                    column_mapping['RTO'] = col_idx
                elif 'Incomplete Address' in header_str:
                    column_mapping['Incomplete Address'] = col_idx
                elif 'Doctor Non Contactable' in header_str or 'Dr. Non contactable' in header_str:
                    column_mapping['Doctor Non Contactable'] = col_idx
                elif 'Doctor Refused' in header_str or 'Refused to Accept' in header_str:
                    column_mapping['Doctor Refused to Accept'] = col_idx
                elif 'Hold Delivery' in header_str:
                    column_mapping['Hold Delivery'] = col_idx
        
        # Clear existing data
        max_clear_rows = max(len(summary_df) + 10, 50)
        for r in range(data_start_row, data_start_row + max_clear_rows):
            for c in range(1, ws.max_column + 1):
                try:
                    ws.cell(row=r, column=c).value = None
                except:
                    pass

        def copy_row_style(src_row_idx, dst_row_idx):
            for c in range(1, ws.max_column + 1):
                try:
                    src = ws.cell(row=src_row_idx, column=c)
                    dst = ws.cell(row=dst_row_idx, column=c)
                    
                    if src.font:
                        dst.font = copy_style(src.font)
                    if src.alignment:
                        dst.alignment = copy_style(src.alignment)
                    if src.border:
                        dst.border = copy_style(src.border)
                    if src.fill:
                        dst.fill = copy_style(src.fill)
                    dst.number_format = src.number_format
                except:
                    pass

        # Write data rows
        template_data_row = data_start_row
        for i in range(len(summary_df)):
            target_row = data_start_row + i
            copy_row_style(template_data_row, target_row)
            
            for col_name, col_idx in column_mapping.items():
                if col_name in summary_df.columns:
                    value = summary_df.iloc[i][col_name]
                    try:
                        cell = ws.cell(row=target_row, column=col_idx)
                        cell.value = value
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            cell.number_format = '0'
                    except:
                        pass

        # total_row = data_start_row + len(summary_df)
        copy_row_style(template_data_row, total_row)
        
        if 'ABM Name' in column_mapping:
            try:
                cell = ws.cell(row=total_row, column=column_mapping['ABM Name'])
                cell.value = "Total"
                cell.font = Font(bold=True, name='Arial', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            except:
                pass
        
        for col_name, col_idx in column_mapping.items():
            if col_name in summary_df.columns and col_name not in ['Area Name', 'ABM Name']:
                total_value = int(summary_df[col_name].sum())
                try:
                    cell = ws.cell(row=total_row, column=col_idx)
                    cell.value = total_value
                    cell.font = Font(bold=True, name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = '0'
                except:
                    pass

        # Save file
        safe_zbm_name = str(zbm_name).replace(' ', '_').replace('/', '_').replace('\\', '_')
        filename = f"ZBM_Summary_{zbm_code}_{safe_zbm_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        wb.save(filepath)
        print(f"   ✅ Created: {filename}")
        
    except Exception as e:
        print(f"   ❌ Error creating Excel report for {zbm_code}: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    create_zbm_hierarchical_reports()
