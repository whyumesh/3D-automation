#!/usr/bin/env python3
"""
ZBM Email Display - Using Pre-generated Reports
Displays professional emails in Outlook for ZBMs using existing summary files
USES OUTLOOK TO DISPLAY EMAILS (NOT SEND AUTOMATICALLY)
"""

import pandas as pd
import os
from datetime import datetime
import warnings
import win32com.client
from openpyxl import load_workbook
import glob

# Suppress pandas warnings
warnings.filterwarnings('ignore', category=FutureWarning, module='pandas')

def send_zbm_emails():
    """Display emails in Outlook for review using pre-generated ZBM summary reports"""
    
    print("🚀 Starting ZBM Email Display (using pre-generated reports)...")
    print("📧 This will DISPLAY emails in Outlook for review - NOT SEND automatically")
    
    # Check if zbm summary folder exists
    zbm_summary_folder = 'zbm summary'
    if not os.path.exists(zbm_summary_folder):
        print(f"❌ Error: '{zbm_summary_folder}' folder not found!")
        print(f"   Please ensure the folder exists with ZBM summary Excel files")
        return
    
    # Find all ZBM summary files
    print(f"📂 Looking for ZBM summary files in '{zbm_summary_folder}' folder...")
    summary_files = glob.glob(os.path.join(zbm_summary_folder, 'ZBM_Summary_*.xlsx'))
    
    if len(summary_files) == 0:
        print(f"❌ No ZBM summary files found in '{zbm_summary_folder}' folder!")
        print(f"   Expected files matching pattern: ZBM_Summary_*.xlsx")
        return
    
    print(f"✅ Found {len(summary_files)} ZBM summary file(s)")
    
    # Read Sample Master Tracker to get ZBM email addresses
    print("📖 Reading Sample Master Tracker.xlsx for email addresses...")
    try:
        df = pd.read_excel('Sample Master Tracker.xlsx')
        print(f"✅ Successfully loaded {len(df)} records from Sample Master Tracker.xlsx")
    except Exception as e:
        print(f"❌ Error reading Sample Master Tracker.xlsx: {e}")
        return
    
    # Get unique ZBM details
    zbm_details = df[['ZBM Terr Code', 'ZBM Name', 'ZBM EMAIL_ID']].drop_duplicates()
    zbm_lookup = {}
    for _, row in zbm_details.iterrows():
        zbm_lookup[row['ZBM Terr Code']] = {
            'name': row['ZBM Name'],
            'email': row['ZBM EMAIL_ID']
        }
    
    # Initialize Outlook
    print("📧 Initializing Outlook...")
    outlook = None
    
    outlook_methods = [
        "Outlook.Application",
        "Outlook.Application.16",
        "Outlook.Application.15",
        "Outlook.Application.14",
        "Outlook.Application.12",
    ]
    
    for method in outlook_methods:
        try:
            print(f"   🔄 Trying: {method}")
            outlook = win32com.client.Dispatch(method)
            print(f"✅ Outlook initialized successfully using: {method}")
            break
        except Exception as e:
            print(f"   ❌ Failed with {method}: {e}")
            continue
    
    if outlook is None:
        print("❌ Could not initialize Outlook with any method")
        print("🔧 Troubleshooting steps:")
        print("   1. Ensure Outlook is installed on this computer")
        print("   2. Try opening Outlook manually first")
        print("   3. Check if Outlook is running in the background")
        print("   4. Try running as administrator")
        print("   5. Install Microsoft Office/Outlook if not present")
        
        print("\n🔄 Creating HTML email files as fallback...")
        create_html_email_files_from_summaries(summary_files, zbm_lookup)
        return
    
    # Process each ZBM summary file
    success_count = 0
    error_count = 0
    
    for summary_file in summary_files:
        filename = os.path.basename(summary_file)
        print(f"\n🔄 Processing: {filename}")
        
        try:
            # Extract ZBM code from filename (format: ZBM_Summary_ZNXX_Name_timestamp.xlsx)
            parts = filename.replace('.xlsx', '').split('_')
            zbm_code = None
            
            # Find the ZN code in the filename
            for part in parts:
                if part.startswith('ZN'):
                    zbm_code = part
                    break
            
            if not zbm_code:
                print(f"   ⚠️ Could not extract ZBM code from filename: {filename}")
                error_count += 1
                continue
            
            # Get ZBM details
            if zbm_code not in zbm_lookup:
                print(f"   ⚠️ ZBM code {zbm_code} not found in Sample Master Tracker")
                error_count += 1
                continue
            
            zbm_name = zbm_lookup[zbm_code]['name']
            zbm_email = zbm_lookup[zbm_code]['email']
            
            print(f"   📋 ZBM: {zbm_code} - {zbm_name}")
            print(f"   📧 Email: {zbm_email}")
            
            # Read the summary file
            summary_df = read_zbm_summary_file(summary_file)
            
            if summary_df is None or len(summary_df) == 0:
                print(f"   ⚠️ No data found in summary file")
                error_count += 1
                continue
            
            print(f"   📊 Found {len(summary_df)} ABMs in summary")
            
            # Get ABM emails for CC
            abm_emails = get_abm_emails_from_tracker(df, zbm_code)
            cc_emails = ', '.join(abm_emails) if abm_emails else ''
            
            # Generate email content
            email_content = generate_email_content(zbm_name, summary_df)
            
            # Display email in Outlook
            display_single_email(outlook, zbm_email, cc_emails, email_content, 
                               zbm_code, zbm_name, summary_file)
            
            success_count += 1
            print(f"   ✅ Email displayed in Outlook for {zbm_name}")
            
        except Exception as e:
            error_count += 1
            print(f"   ❌ Error processing {filename}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    print(f"\n🎉 Email display completed!")
    print(f"✅ Successfully displayed: {success_count} emails")
    print(f"❌ Failed to display: {error_count} emails")
    print(f"\n📧 All emails are now open in Outlook for your review and manual sending")

def read_zbm_summary_file(filepath):
    """Read ZBM summary file and extract data"""
    
    try:
        # Load workbook
        wb = load_workbook(filepath, data_only=True)
        
        # Try to find the ZBM sheet
        sheet_name = None
        if 'ZBM' in wb.sheetnames:
            sheet_name = 'ZBM'
        else:
            sheet_name = wb.sheetnames[0]
        
        ws = wb[sheet_name]
        
        # Find the header row (should be row 7 based on template)
        header_row = 7
        
        # Read headers from row 7
        headers = []
        for col in range(5, 23):  # Columns E to V
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"Column_{col}")
        
        # Read data rows (starting from row 8)
        data = []
        row_num = 8
        
        while True:
            # Check if this is the total row or empty row
            first_col = ws.cell(row=row_num, column=6).value  # ABM Name column
            
            if first_col is None or str(first_col).strip().lower() == 'total':
                break
            
            # Read row data
            row_data = {}
            for idx, col in enumerate(range(5, 23)):
                cell_value = ws.cell(row=row_num, column=col).value
                row_data[headers[idx]] = cell_value
            
            data.append(row_data)
            row_num += 1
            
            # Safety check
            if row_num > 1000:
                break
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Clean column names
        df.columns = df.columns.str.replace('\n', ' ').str.strip()
        
        return df
        
    except Exception as e:
        print(f"      ❌ Error reading summary file: {e}")
        return None

def get_abm_emails_from_tracker(df, zbm_code):
    """Get ABM email addresses for a specific ZBM from the master tracker"""
    
    try:
        zbm_data = df[df['ZBM Terr Code'] == zbm_code]
        abm_emails = zbm_data['ABM EMAIL_ID'].dropna().unique().tolist()
        return [email for email in abm_emails if email and str(email).strip()]
    except Exception as e:
        print(f"      ⚠️ Could not get ABM emails: {e}")
        return []

def generate_email_content(zbm_name, summary_df):
    """Generate professional email content"""
    
    current_date = datetime.now().strftime('%B %d, %Y')
    
    # Create summary table HTML
    table_html = create_summary_table_html(summary_df)
    
    email_content = f"""
Hi {zbm_name},

Please refer the status Sample requests raised in Abbworld for your area.

{table_html}

You can track your sample request at the following link with the Docket Number:

DTDC: <a href="https://www.dtdc.com/tracking">Click here</a>

Speed Post: <a href="https://www.indiapost.gov.in/vas/Pages/IndiaPostHome.aspx">Click Here</a>

In case of any query, please contact 1Point.

Regards,
Umesh Pawar.
"""
    
    return email_content

def create_summary_table_html(summary_df):
    """Create HTML table for summary data"""
    
    if summary_df.empty:
        return "<p>No data available</p>"
    
    # Create HTML table
    html = "<table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%; font-size: 12px;'>"
    
    # Header row
    html += "<tr style='background-color: #f0f0f0; font-weight: bold;'>"
    for col in summary_df.columns:
        # Format column headers
        col_display = str(col).replace('\n', '<br/>')
        html += f"<th>{col_display}</th>"
    html += "</tr>"
    
    # Data rows
    for _, row in summary_df.iterrows():
        html += "<tr>"
        for col in summary_df.columns:
            value = row[col]
            # Format value
            if pd.isna(value):
                value = 0
            elif isinstance(value, (int, float)):
                value = int(value)
            html += f"<td>{value}</td>"
        html += "</tr>"
    
    # Total row
    html += "<tr style='background-color: #e0e0e0; font-weight: bold;'>"
    for idx, col in enumerate(summary_df.columns):
        if idx == 0:
            html += "<td>TOTAL</td>"
        elif idx == 1:
            html += "<td></td>"
        else:
            try:
                total = summary_df[col].sum()
                html += f"<td>{int(total)}</td>"
            except:
                html += "<td></td>"
    html += "</tr>"
    
    html += "</table>"
    
    return html

def display_single_email(outlook, zbm_email, cc_emails, email_content, 
                        zbm_code, zbm_name, summary_file):
    """Display a single email in Outlook for review (without sending)"""
    
    # Create new mail item
    mail = outlook.CreateItem(0)  # 0 = olMailItem
    
    # Set recipients
    mail.To = zbm_email
    
    # Set CC recipients
    if cc_emails:
        mail.CC = cc_emails
    
    # Set subject
    current_date = datetime.now().strftime('%B %d, %Y')
    mail.Subject = f"Sample Direct Dispatch to Doctors - Request Status as of {current_date}"
    
    # Set body
    mail.HTMLBody = email_content
    
    # Add attachment (the summary file itself)
    if summary_file and os.path.exists(summary_file):
        mail.Attachments.Add(os.path.abspath(summary_file))
        print(f"   📎 Attached: {os.path.basename(summary_file)}")
    
    # Display email (don't send)
    mail.Display()
    
    print(f"   📧 Email displayed for: {zbm_email}")
    if cc_emails:
        print(f"   📧 CC'd to: {cc_emails}")
    print(f"   ⚠️  Review the email and send manually from Outlook")

def create_html_email_files_from_summaries(summary_files, zbm_lookup):
    """Create HTML email files as fallback when Outlook is not available"""
    
    print("📧 Creating HTML email files...")
    
    # Create output directory
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = f"ZBM_HTML_Emails_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Created output directory: {output_dir}")
    
    success_count = 0
    
    for summary_file in summary_files:
        filename = os.path.basename(summary_file)
        print(f"\n🔄 Processing: {filename}")
        
        try:
            # Extract ZBM code
            parts = filename.replace('.xlsx', '').split('_')
            zbm_code = None
            for part in parts:
                if part.startswith('ZN'):
                    zbm_code = part
                    break
            
            if not zbm_code or zbm_code not in zbm_lookup:
                print(f"   ⚠️ Skipping {filename}")
                continue
            
            zbm_name = zbm_lookup[zbm_code]['name']
            zbm_email = zbm_lookup[zbm_code]['email']
            
            # Read summary
            summary_df = read_zbm_summary_file(summary_file)
            if summary_df is None or len(summary_df) == 0:
                continue
            
            # Generate email content
            email_content = generate_email_content(zbm_name, summary_df)
            
            # Create HTML file
            create_single_html_email(zbm_code, zbm_name, zbm_email, '', 
                                    email_content, output_dir)
            
            success_count += 1
            print(f"   ✅ HTML email created for {zbm_name}")
            
        except Exception as e:
            print(f"   ❌ Error: {e}")
            continue
    
    print(f"\n🎉 HTML email creation completed!")
    print(f"✅ Successfully created: {success_count} HTML email files")
    print(f"📁 Files saved in: {output_dir}")

def create_single_html_email(zbm_code, zbm_name, zbm_email, cc_emails, 
                            email_content, output_dir):
    """Create a single HTML email file"""
    
    current_date = datetime.now().strftime('%B %d, %Y')
    
    html_email = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Sample Direct Dispatch to Doctors - Request Status as of {current_date}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; font-weight: bold; }}
        .total-row {{ background-color: #e0e0e0; font-weight: bold; }}
        .header {{ background-color: #f0f0f0; padding: 10px; margin-bottom: 20px; }}
    </style>
</head>
<body>
    <div class="header">
        <h3>Email Details:</h3>
        <p><strong>To:</strong> {zbm_email}</p>
        <p><strong>CC:</strong> {cc_emails}</p>
        <p><strong>Subject:</strong> Sample Direct Dispatch to Doctors - Request Status as of {current_date}</p>
    </div>
    
    <div class="email-content">
        {email_content}
    </div>
</body>
</html>
"""
    
    # Save HTML file
    safe_zbm_name = str(zbm_name).replace(' ', '_').replace('/', '_').replace('\\', '_')
    filename = f"Email_{zbm_code}_{safe_zbm_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    filepath = os.path.join(output_dir, filename)
    
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html_email)
    
    print(f"   📧 HTML email saved: {filename}")

if __name__ == "__main__":
    send_zbm_emails()
